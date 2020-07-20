
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment,colors
import xlsxwriter
from openpyxl.drawing.image import Image
from win32com import client
import win32com.client as w3c
import win32api
import time
import os,sys









wb = openpyxl.load_workbook('New_Salary.xlsx',data_only=True)
wb2= openpyxl.load_workbook('Salary.xlsx')

cwd = os.getcwd()
# exce_file=os.mkdir("Excel_files")
# pdf_files=os.mkdir("Pdf_Files")


sheet=wb.sheetnames
ws=wb[sheet[0]]
sheet2=wb2.sheetnames
ws2=wb2[sheet2[0]]


thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))


thick_border= Border(left=Side(style='thick'),
                         right=Side(style='thick'),
                         top=Side(style='thick'),
                         bottom=Side(style='thick'))





rows=list(ws.iter_rows())
for row in rows[1:]:
    filename = row[0].value
    wb3=openpyxl.load_workbook("Salary.xlsx",data_only=True)
    sheet3 = wb3.sheetnames
    ws3 = wb3[sheet3[0]]




    ws3.cell(row=4, column=3).value = row[0].value
    # print(row[3].value)
    # print(row[38].value)
    ws3.cell(row=4, column=7).value = row[25].value
    ws3.cell(row=5, column=3).value = row[22].value
    ws3.cell(row=5, column=7).value = row[26].value
    ws3.cell(row=6,column=3).value=row[23].value
    ws3.cell(row=6,column=7).value=row[27].value
    ws3.cell(row=7,column=3).value=row[24].value
    ws3.cell(row=9,column=2).value=row[28].value
    ws3.cell(row=9,column=4).value=row[29].value
    ws3.cell(row=9,column=6).value=row[30].value
    ws3.cell(row=11,column=2).value=row[31].value
    ws3.cell(row=11,column=5).value=row[32].value
    ws3.cell(row=11,column=7).value=row[33].value
    ws3.cell(row=13,column=2).value=row[34].value
    ws3.cell(row=13,column=3).value=row[35].value
    ws3.cell(row=13, column=5).value = row[36].value
    ws3.cell(row=13, column=7).value =row[37].value
    ws3.cell(row=15, column=4).value = row[3].value
    ws3.cell(row=16, column=4).value = row[4].value
    ws3.cell(row=17, column=4).value = row[5].value
    ws3.cell(row=18, column=4).value = row[6].value
    ws3.cell(row=19, column=4).value = row[7].value
    ws3.cell(row=20, column=4).value = row[8].value
    ws3.cell(row=21, column=4).value = row[10].value
    ws3.cell(row=22, column=4).value = row[11].value
    ws3.cell(row=23, column=4).value = row[38].value
    ws3.cell(row=24, column=4).value = row[12].value
    ws3.cell(row=15, column=8).value = row[15].value
    ws3.cell(row=16, column=8).value = row[16].value
    ws3.cell(row=17, column=8).value = row[13].value
    ws3.cell(row=18, column=8).value = row[14].value
    ws3.cell(row=23, column=8).value = row[39].value



    # img = openpyxl.drawing.image.Image('test.jpg')
    # ws3.add_image(img, 'H1')


    for row in range(2,28):
        for column in range(2,10):
            ws3.cell(row=row,column=column).border=thin_border




    wb3.save( "C:\Python27\Scripts\\Salary\Excel_Files\\" + filename + ".xlsx")




    xlApp = w3c.gencache.EnsureDispatch("Excel.Application")
    print(filename)
    books = xlApp.Workbooks.Open("C:\Python27\Scripts\\Salary\Excel_Files\\" + filename)
    xlsheet = books.Worksheets('sheet1')
    ws = books.Worksheets[0]
    ws.Visible = 1
    ws.PageSetup.FitToPagesTall = 1
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.Zoom = False
    ws.PageSetup.PrintArea = 'B1:I27'
    ws.ExportAsFixedFormat(0, 'C:\Python27\Scripts\\Salary\PDF_Files\\'+ filename + ".pdf")



# o = win32com.client.Dispatch("Excel.Application")
#
# o.Visible = False
#
# wb_path = r'c:\user\desktop\sample.xls'
#
# wb = o.Workbooks.Open(wb_path)
#
#
#
# ws_index_list = [1,4,5] #say you want to print these sheets
#
# path_to_pdf = r'C:\user\desktop\sample.pdf'
#
#
#
# wb.WorkSheets(ws_index_list).Select()
#
# wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)











































# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A2'] = 'Tom'
# ws['B2'] = 30
#
# ws['A3'] = 'Marry'
# ws['B3'] = 29
#
# # Save the file
# wb.save("new.xlsx")